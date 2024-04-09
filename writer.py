from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


#adder -> +5 each time
#n = group number  cell a1 to a5

#name = group name  cell b1 to b 5
#list1 = registeration numbers  cell d1,d2,d3,d4,d5

def write(n,membercount,teamnum,teamname,members):
    try:
        # Try to load the workbook
        wb = load_workbook("teams.xlsx")
        ws = wb['Sheet1']
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        print("Error: The file 'teams.xlsx' does not exist. Creating a new workbook.")
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        # Perform operations here
        print("New workbook created.")
    except Exception as e:
        print("An error occurred:", e)
    
    
    try:
        Astart = "A" + str(n+1)
        Aend = "A"+str(n+membercount)
        Bstart = "B" + str(n+1)
        Bend = "B"+str(n+membercount)
        ws.merge_cells(f"{Astart}:{Aend}")
        ws.merge_cells(f"{Bstart}:{Bend}")
        ws[Astart].alignment = Alignment(horizontal = "center",vertical = "center")
        ws[Bstart].alignment = Alignment(horizontal = "center",vertical = "center")
        ws[Astart] = teamnum
        ws[Bstart] = teamname
        count = n

        for member in members:
            count+=1
            name,email = findname(member)
            ws["C"+str(count)]=member
            ws["D"+str(count)]=name
            ws["E"+str(count)]=email
        wb.save("teams.xlsx")
        print("Operations completed successfully.")
        return n+membercount
    except Exception as e:
        print("An error occurred during operations:", e)
    




def load_data(filename):
    data = {}
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        reg_num = row[0]
        name = row[1]
        email = row[2]
        data[reg_num] = (name, email)
    workbook.close()
    return data


data_dict = load_data('data.xlsx')

# Function to find name and email from registration number using the dictionary
def findname(reg):
    return data_dict.get(reg, (None, None))
        
        
        
        
        
        
        
        
# def findname(reg):
#         workbook1 = load_workbook('data.xlsx')
#         sheet = workbook1.active

#         # Find the row with the registration number
#         for row in sheet.iter_rows(values_only=True):
#             if row[0] == reg:
#                 email = row[2]
#                 name = row[1]
#                 return name,email
#         return None
                
        
        
