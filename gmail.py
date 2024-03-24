import openpyxl
import random
import string
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
import otp as o



def Name(reg_no):
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active

    # Find the row with the registration number
    for row in sheet.iter_rows(values_only=True):
        if row[0] == reg_no:
            email = row[2]
            name = row[1]
    return name

def send_otp(reg_no):
    # Load the Excel sheet
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook.active

    # Find the row with the registration number
    for row in sheet.iter_rows(values_only=True):
        if row[0] == reg_no:
            email = row[2]
            name = row[1]
            break
    else:
        return None  # Registration number not found

    # Generate a unique OTP
    otp = ''.join(random.choices(string.digits, k=8))

    # Craft the HTML message
    html_message = f"""
  <!DOCTYPE html>
  <html lang="en">
  <head>
    <meta charset="UTF-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Welcome to ExhibitionEase - Verify Your Account</title>
    <style>
      body {{
        font-family: Arial, sans-serif;
        margin: 0;
        padding: 0;
        color: #333; /* Darker text for better readability */
        line-height: 1.5;
      }}

      p {{ margin: 15px 0; }}

      h1, h2 {{
        margin: 20px 0;
        color: #29ABCA; /* Brand blue for headings */
      }}

      a {{ color: #29ABCA; /* Brand blue for links */
        text-decoration: none;
      }}
      a:hover {{ text-decoration: underline; }}

      .container {{
        max-width: 600px;
        margin: 0 auto;
        padding: 30px;
        background-color: #F8F9F9; /* Lighter background for contrast */
      }}

      .header {{
        text-align: center;
      }}

      .footer {{
        text-align: center;
        margin-top: 20px;
        color: #aaa;
      }}
    </style>
  </head>
  <body>
    <div class="container">
      <div class="header">
        <h1>Welcome to ExhibitionEase, {name}!</h1>
      </div>

      <p>Thank you for choosing ExhibitionEase. We're excited to have you on board!</p>

      <p>To ensure the security of your account, please use the following One-Time Password (OTP) to complete the verification process:</p>

      <h2 style="text-align: center; background-color: #E0E0E0; padding: 10px; border-radius: 5px;">OTP: {otp}</h2>

      <p>Your trust in us is valued. We're committed to providing exceptional support. If you have any questions or need assistance, please don't hesitate to contact our dedicated team. We're here to help you every step of the way.</p>

      <p>We look forward to making your exhibition experience a breeze!</p>

      <p>Best regards,</p>

      <p>The ExhibitionEase Team</p>

      <div class="footer">
        <p>&copy; 2024 ExhibitionEase. All rights reserved.</p>
      </div>
    </div>
  </body>
  </html>
  """

    # Create message container - the correct MIME type is multipart/alternative.
    msg = MIMEMultipart('alternative')
    msg['Subject'] = 'OTP Verification'
    msg['From'] = "exhibitionease.auth@gmail.com"
    msg['To'] = email

    # Record the MIME types of both parts - text/plain and text/html.
    part1 = MIMEText(html_message, 'html')

    # Attach parts into message container.
    msg.attach(part1)

    # Send the email
    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login("exhibitionease.auth@gmail.com", os.getenv('email_key'))
        server.sendmail("exhibitionease.auth@gmail.com", email, msg.as_string())

    return otp
    # Store the OTP in the temporary directory



